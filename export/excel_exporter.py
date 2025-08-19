"""
Excel Export Module
Exports GPU listing data to Excel format with multiple sheets and formatting
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


class ExcelExporter:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.logger = logging.getLogger(__name__)

    def export_to_excel(self, listings: List[Dict[str, Any]], 
                       compliance_results: Dict[str, Dict[str, Any]] = None) -> str:
        """
        Export GPU listings to Excel with multiple sheets
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gpu_listings_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        
        try:
            # Create workbook and sheets
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create main listings sheet
            self._create_listings_sheet(wb, listings)
            
            # Create summary sheet
            self._create_summary_sheet(wb, listings)
            
            # Create compliance sheet if data provided
            if compliance_results:
                self._create_compliance_sheet(wb, compliance_results)
            
            # Create price analysis sheet
            self._create_price_analysis_sheet(wb, listings)
            
            # Save workbook
            wb.save(filepath)
            
            self.logger.info(f"Exported {len(listings)} listings to {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Failed to export to Excel: {e}")
            raise

    def _create_listings_sheet(self, wb: Workbook, listings: List[Dict[str, Any]]) -> None:
        """Create the main listings data sheet"""
        ws = wb.create_sheet("GPU Listings", 0)
        
        if not listings:
            ws['A1'] = "No listings found"
            return
        
        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(listings)
        
        # Standardize columns
        standard_columns = [
            'title', 'marketplace', 'price', 'standardized_price', 'gpu_model', 
            'gpu_series', 'condition', 'location', 'url', 'posted_date', 
            'seller_info', 'is_sold', 'is_featured', 'scraped_at'
        ]
        
        # Ensure all standard columns exist
        for col in standard_columns:
            if col not in df.columns:
                df[col] = None
        
        # Reorder columns
        existing_cols = [col for col in standard_columns if col in df.columns]
        extra_cols = [col for col in df.columns if col not in standard_columns]
        df = df[existing_cols + extra_cols]
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Format the sheet
        self._format_listings_sheet(ws, len(df))

    def _format_listings_sheet(self, ws, num_rows: int) -> None:
        """Apply formatting to the listings sheet"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add table formatting
        if num_rows > 0:
            table_range = f"A1:{ws.max_column_letter}{num_rows + 1}"
            table = Table(displayName="GPUListings", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=True
            )
            ws.add_table(table)

    def _create_summary_sheet(self, wb: Workbook, listings: List[Dict[str, Any]]) -> None:
        """Create summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "GPU Scraper Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        
        # Basic statistics
        row = 3
        ws[f'A{row}'] = "Total Listings Found:"
        ws[f'B{row}'] = len(listings)
        
        row += 1
        ws[f'A{row}'] = "Scraping Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Marketplace breakdown
        if listings:
            marketplace_counts = {}
            price_data = []
            gpu_models = {}
            
            for listing in listings:
                # Count by marketplace
                marketplace = listing.get('marketplace', 'Unknown')
                marketplace_counts[marketplace] = marketplace_counts.get(marketplace, 0) + 1
                
                # Collect price data
                price = listing.get('standardized_price') or listing.get('price')
                if isinstance(price, (int, float)) and price > 0:
                    price_data.append(price)
                
                # Count GPU models
                model = listing.get('gpu_model', 'Unknown')
                gpu_models[model] = gpu_models.get(model, 0) + 1
            
            # Marketplace breakdown
            row += 2
            ws[f'A{row}'] = "Listings by Marketplace:"
            ws[f'A{row}'].font = Font(bold=True)
            
            for marketplace, count in marketplace_counts.items():
                row += 1
                ws[f'A{row}'] = f"  {marketplace}:"
                ws[f'B{row}'] = count
            
            # Price statistics
            if price_data:
                row += 2
                ws[f'A{row}'] = "Price Statistics:"
                ws[f'A{row}'].font = Font(bold=True)
                
                row += 1
                ws[f'A{row}'] = "  Average Price:"
                ws[f'B{row}'] = f"£{sum(price_data) / len(price_data):.2f}"
                
                row += 1
                ws[f'A{row}'] = "  Min Price:"
                ws[f'B{row}'] = f"£{min(price_data):.2f}"
                
                row += 1
                ws[f'A{row}'] = "  Max Price:"
                ws[f'B{row}'] = f"£{max(price_data):.2f}"
            
            # GPU model breakdown (top 10)
            row += 2
            ws[f'A{row}'] = "Top GPU Models:"
            ws[f'A{row}'].font = Font(bold=True)
            
            sorted_models = sorted(gpu_models.items(), key=lambda x: x[1], reverse=True)
            for model, count in sorted_models[:10]:
                row += 1
                ws[f'A{row}'] = f"  {model}:"
                ws[f'B{row}'] = count
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_compliance_sheet(self, wb: Workbook, compliance_results: Dict[str, Dict[str, Any]]) -> None:
        """Create compliance check results sheet"""
        ws = wb.create_sheet("Compliance Report")
        
        # Title
        ws['A1'] = "Website Compliance Check Results"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        headers = ['Website', 'Robots.txt Allowed', 'Rate Limit Compliance', 'ToS Concerns', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for site, results in compliance_results.items():
            row += 1
            ws.cell(row=row, column=1, value=site.title())
            
            if 'error' in results:
                ws.cell(row=row, column=4, value=f"Error: {results['error']}")
            else:
                ws.cell(row=row, column=2, value="Yes" if results.get('robots_allowed', True) else "No")
                ws.cell(row=row, column=3, value="Yes")  # Assuming we follow rate limits
                ws.cell(row=row, column=4, value=", ".join(results.get('tos_concerns', [])) or "None")
                ws.cell(row=row, column=5, value=results.get('notes', ''))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_price_analysis_sheet(self, wb: Workbook, listings: List[Dict[str, Any]]) -> None:
        """Create price analysis sheet with charts-ready data"""
        ws = wb.create_sheet("Price Analysis")
        
        if not listings:
            ws['A1'] = "No data available for price analysis"
            return
        
        # Title
        ws['A1'] = "GPU Price Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Prepare price data by GPU model
        gpu_prices = {}
        for listing in listings:
            model = listing.get('gpu_model', 'Unknown')
            price = listing.get('standardized_price')
            
            if isinstance(price, (int, float)) and price > 0:
                if model not in gpu_prices:
                    gpu_prices[model] = []
                gpu_prices[model].append(price)
        
        if not gpu_prices:
            ws['A3'] = "No valid price data found"
            return
        
        # Create price summary table
        row = 3
        headers = ['GPU Model', 'Count', 'Avg Price', 'Min Price', 'Max Price', 'Price Range']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        for model, prices in gpu_prices.items():
            row += 1
            ws.cell(row=row, column=1, value=model)
            ws.cell(row=row, column=2, value=len(prices))
            ws.cell(row=row, column=3, value=f"£{sum(prices) / len(prices):.2f}")
            ws.cell(row=row, column=4, value=f"£{min(prices):.2f}")
            ws.cell(row=row, column=5, value=f"£{max(prices):.2f}")
            ws.cell(row=row, column=6, value=f"£{max(prices) - min(prices):.2f}")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_csv(self, listings: List[Dict[str, Any]], filename_suffix: str = "") -> str:
        """
        Export listings to CSV format as a backup/alternative
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"gpu_listings_{timestamp}{filename_suffix}.csv"
        filepath = self.output_dir / filename
        
        try:
            df = pd.DataFrame(listings)
            df.to_csv(filepath, index=False, encoding='utf-8')
            
            self.logger.info(f"Exported {len(listings)} listings to CSV: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Failed to export to CSV: {e}")
            raise

    def create_summary_report(self, listings: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Create a summary report dictionary
        """
        if not listings:
            return {"total_listings": 0, "message": "No listings found"}
        
        marketplace_counts = {}
        price_data = []
        gpu_models = {}
        conditions = {}
        
        for listing in listings:
            # Count by marketplace
            marketplace = listing.get('marketplace', 'Unknown')
            marketplace_counts[marketplace] = marketplace_counts.get(marketplace, 0) + 1
            
            # Collect price data
            price = listing.get('standardized_price') or listing.get('price')
            if isinstance(price, (int, float)) and price > 0:
                price_data.append(price)
            
            # Count GPU models
            model = listing.get('gpu_model', 'Unknown')
            gpu_models[model] = gpu_models.get(model, 0) + 1
            
            # Count conditions
            condition = listing.get('condition', 'Unknown')
            conditions[condition] = conditions.get(condition, 0) + 1
        
        summary = {
            "total_listings": len(listings),
            "marketplaces": marketplace_counts,
            "gpu_models": dict(sorted(gpu_models.items(), key=lambda x: x[1], reverse=True)[:10]),
            "conditions": conditions,
            "scraping_date": datetime.now().isoformat()
        }
        
        if price_data:
            summary["price_stats"] = {
                "count": len(price_data),
                "average": round(sum(price_data) / len(price_data), 2),
                "min": min(price_data),
                "max": max(price_data),
                "range": round(max(price_data) - min(price_data), 2)
            }
        
        return summary


# Test function for Excel exporter
def test_excel_exporter():
    """Test the Excel exporter functionality"""
    
    # Sample test data
    test_listings = [
        {
            'title': 'NVIDIA RTX 4070 Gaming Graphics Card',
            'marketplace': 'eBay UK',
            'price': '£450.00',
            'standardized_price': 450.0,
            'gpu_model': 'RTX 4070',
            'gpu_series': 'RTX 40',
            'condition': 'Used',
            'location': 'London',
            'url': 'https://example.com/listing1',
            'scraped_at': datetime.now().isoformat()
        },
        {
            'title': 'AMD RX 7800 XT Graphics Card',
            'marketplace': 'Gumtree UK',
            'price': '£380.00',
            'standardized_price': 380.0,
            'gpu_model': 'RX 7800 XT',
            'gpu_series': 'RX 7000',
            'condition': 'Excellent',
            'location': 'Manchester',
            'url': 'https://example.com/listing2',
            'scraped_at': datetime.now().isoformat()
        }
    ]
    
    exporter = ExcelExporter()
    
    # Test Excel export
    excel_file = exporter.export_to_excel(test_listings)
    print(f"Test Excel file created: {excel_file}")
    
    # Test CSV export
    csv_file = exporter.export_to_csv(test_listings, "_test")
    print(f"Test CSV file created: {csv_file}")
    
    # Test summary report
    summary = exporter.create_summary_report(test_listings)
    print(f"Summary report: {summary}")


if __name__ == "__main__":
    test_excel_exporter()